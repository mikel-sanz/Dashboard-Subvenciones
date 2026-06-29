"""
🛡️ NEXO MDU: test_reports.py
Propósito: Pruebas unitarias para el exportador de informes PDF y Excel.
Arquitectura: Modular Design Unit (MDU) - Suite de Pruebas Unitarias.
"""

import datetime
import io

import openpyxl
import pandas as pd

from src.visualization.reports import ReportExporter


def _crear_df_prueba() -> pd.DataFrame:
    """Genera un DataFrame de prueba con datos de subvenciones."""
    return pd.DataFrame({
        "Tipo_Subvencion": [
            "Ayuda A", "Subvención B", "Programa C",
        ],
        "Cuantia": [50000.0, 120000.0, 75000.0],
        "Fecha_Vigencia": [
            datetime.date(2026, 12, 31),
            datetime.date(2026, 6, 30),
            datetime.date(2027, 3, 15),
        ],
        "Entidad_Convocante": [
            "Gobierno de Navarra",
            "Ministerio de Industria",
            "Comisión Europea",
        ],
        "Ambito_Territorial": [
            "Navarra", "España", "Europa",
        ],
        "Actividad_Relacionada": [
            "Digitalización/Robótica",
            "Transición Verde/Sostenibilidad",
            "I+D+i Científica",
        ],
        "URL_Convocatoria": [
            "https://example.com/a",
            "https://example.com/b",
            "https://example.com/c",
        ],
    })


def test_generar_excel_retorna_bytes_validos() -> None:
    """
    ARRANGE: Crear un DataFrame con 3 registros de prueba.
    ACT: Generar el archivo Excel en memoria.
    ASSERT: Verificar que retorna bytes no vacíos y que se
            puede abrir con openpyxl sin errores.
    """
    # Arrange
    df = _crear_df_prueba()

    # Act
    resultado = ReportExporter.generar_excel(df)

    # Assert
    assert isinstance(resultado, bytes)
    assert len(resultado) > 0

    # Verificar que se puede abrir como un workbook válido
    wb = openpyxl.load_workbook(io.BytesIO(resultado))
    assert "Resumen KPI" in wb.sheetnames
    assert "Convocatorias" in wb.sheetnames
    # Verificar que hay datos en la hoja de convocatorias
    ws = wb["Convocatorias"]
    assert ws.max_row >= 4  # cabecera + 3 filas


def test_generar_pdf_retorna_bytes_validos() -> None:
    """
    ARRANGE: Crear un DataFrame con 3 registros de prueba.
    ACT: Generar el archivo PDF en memoria.
    ASSERT: Verificar que retorna bytes que comienzan con '%PDF-'.
    """
    # Arrange
    df = _crear_df_prueba()

    # Act
    resultado = ReportExporter.generar_pdf(df)

    # Assert
    assert isinstance(resultado, bytes)
    assert len(resultado) > 0
    assert resultado[:5] == b"%PDF-"


def test_generar_informes_con_dataframe_vacio() -> None:
    """
    ARRANGE: Crear un DataFrame vacío.
    ACT: Generar Excel y PDF.
    ASSERT: Verificar que no lanzan excepciones y retornan bytes.
    """
    # Arrange
    df = pd.DataFrame()

    # Act
    excel = ReportExporter.generar_excel(df)
    pdf = ReportExporter.generar_pdf(df)

    # Assert
    assert isinstance(excel, bytes) and len(excel) > 0
    assert isinstance(pdf, bytes) and len(pdf) > 0
